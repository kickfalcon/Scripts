import openpyxl as px
import readjson as rj
from dotenv import load_dotenv
import os

# File location

path = os.getenv('FILE_PATH')
file_name = os.getenv('FILE_NAME')
file_path = f"{path}\\{file_name}"

# API URL and headers
# Load environment variables from .env file
load_dotenv()

# Define the base URL and endpoint for the API
base_url = os.getenv('BASE_URL')
endpoint = '/ic/api/integration/v1/integrations'
headers = {
    "Content-Type": "application/json",
    # "Authorization":"Bearer access_token"
}
payload = {
    "q": {
        "status":"ACTIVATED"
    }
}
# Construct the full URL for the API request

url = f"{base_url}{endpoint}"
integrations_json_data = rj.connect_to_api(url, payload=payload, headers=headers)

# Setting the info
rj.connect_to_api()
excel_file = px.load_workbook(file_path)
sheet = excel_file.active

# Generating a new sheet if it does not exist

if "ReportSheet" in excel_file.sheetnames:
    sheet2 = excel_file["ReportSheet"]
else:
    sheet2 = excel_file.create_sheet("ReportSheet")


# Setting styles

yellow_fill = px.styles.PatternFill(
    start_color="FFFF00", 
    end_color="FFFF00",
    fill_type="solid"
)
alert_fill = px.styles.PatternFill(
    start_color="FF0000", 
    end_color="FF0000", 
    fill_type="solid"
)
# Getting the sheet by name
# sheet = excel_file.get_sheet_by_name('Sheet1')

# Getting the data from the first row of the sheet
for row in sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row):
    for cell in row:
        # Getting the integration versions
        gen2_integration = sheet.cell(row=cell.row, column=2).value
        integration_name = sheet.cell(row=cell.row, column=1).value
        gen3_integration = rj.processingJson(integration_name)
        try:
            if gen3_integration is None:
                sheet2.cell(row=cell.row, column=1).value = integration_name
                sheet2.cell(row=cell.row, column=2).value = gen2_integration
            else:
                latest_version = max(gen2_integration, gen3_integration)
                if latest_version == gen3_integration:
                    sheet2.cell(row=cell.row, column=1).value = integration_name
                    sheet2.cell(row=cell.row, column=2).value = gen3_integration
                    sheet2.cell(row=cell.row, column=2).fill = yellow_fill
                else:
                    sheet2.cell(row=cell.row, column=1).value = integration_name
                    sheet2.cell(row=cell.row, column=2).value = gen2_integration
        except TypeError:
            sheet2.cell(row=cell.row, column=1).value = integration_name
            sheet2.cell(row=cell.row, column=2).value = (
                "Null or non-comparable values found."
            )
            sheet2.cell(row=cell.row, column=2).fill = alert_fill
            continue

excel_file.save(file_path)
